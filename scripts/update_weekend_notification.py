#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Aug 30 12:13:07 2018

@author: sagar
"""


import re
import pandas as pd
from openpyxl import load_workbook
from aux_functions import this_weekend, read_sheet
from notifications import instantiate, create_event

google_service = instantiate()

# Contact List
contacts = pd.read_csv("source_data/ContactList.csv")
contacts['Contact#'] = contacts['Contact#'].fillna(0).astype(int).astype(str)

# Program Schedule
program_schedule = load_workbook("source_data/Program Schedule.xlsx",data_only=True)
ps_schema = ["Week Day","Day","Date","Course #","Topics","Mentor","ROTe","CUTe","Topics"]
           
# List of program schedules - 1 dataframe per batch                  
schedule_dict = {}
for sheet_name in program_schedule.sheetnames:
    sheet = read_sheet(program_schedule,sheet_name)
    if sheet.shape[1] >= 9:
        sheet = sheet.iloc[:,0:9]
        if False in set(ps_schema == sheet.columns):
            raise Exception('Column schema mismatch in sheet = ' + sheet_name)
        else:
            schedule_dict[re.findall(r'\d+', sheet_name)[0]] = sheet


# Scheduling Matrix
scheduling_matrix = load_workbook("source_data/Faculty-Scientist_Scheduling_Matrix.xlsx",data_only=True)
ds_schedule = read_sheet(scheduling_matrix,'DS_PGP2018')
ds_schedule = ds_schedule.replace({"":None," ":None,"  ":None,"   ":None})
ds_schedule['Date'] = pd.to_datetime(ds_schedule['Date']).dt.date
# This week's schedule
ds_weekend_schedule = ds_schedule[ds_schedule['Date'].isin(this_weekend())]
#ds_weekend_schedule = ds_schedule[ds_schedule['Date'].isin([day])]

mentor_schedule = read_sheet(scheduling_matrix,'Mentor_PGP2018')
mentor_schedule = mentor_schedule.replace({"":None," ":None,"  ":None,"   ":None})
mentor_schedule['Date'] = pd.to_datetime(mentor_schedule['Date']).dt.date
# This week's schedule
mentor_weekend_schedule = mentor_schedule[mentor_schedule['Date'].isin(this_weekend())]
#mentor_weekend_schedule = mentor_schedule[mentor_schedule['Date'].isin([day])]



# Finding Active batches this week
active_batches = []
for index, row in ds_weekend_schedule.iloc[:,ds_schedule.columns != "Date"].iterrows():
    idx = row.isin(active_batches)
    active_batches.extend(set(row[~idx]))
active_batches.remove(None)


## Batch specifuc notifications
# =============================================================================
# HYD = []
# BLR = []
# for batch in active_batches:
#     if 'HYD' in batch:
#         HYD.append(batch)
#     else:
#         BLR.append(batch)
# 
# active_batches = BLR
# 
# =============================================================================

event_frame = pd.DataFrame(columns=['dsEventId','mentorEventId','Date','Batch','Module','Topic','Mentor','MentorEventType','DSEventType', 'MentorEmail','AttendeeList','MentorDetails','DSContact','ROTe','ROTeDay','ROTeTopics'])
#event_frame = pd.read_csv('/home/sagar/Dropbox/Scheduler/04-Aug_Notifications_bak.csv', index_col = 0)


## Manual Test
#batch = "PGP47-BLR"
#day = this_weekend()[1]

#run = 'dry'
run = 'send'

for batch in active_batches:
    for day in this_weekend():
        print(batch,day)

        ## Data Scientist
        ds_list = ds_weekend_schedule.columns[(ds_weekend_schedule[ds_weekend_schedule['Date']==day]==batch).all()].tolist()
        ds_dict = {}
        for ds in ds_list:
            if (ds in contacts['Mapping'].tolist() or ds in contacts['Name'].tolist()):
                ds_dict[ds] = contacts.loc[(contacts['Mapping']==ds) | (contacts['Name']==ds),['emailid','Contact#']].values[0]
        ds_contacts = '\n'.join(['%s - %s - %s' % (key, value[0], value[1]) for (key, value) in ds_dict.items()])

        ## Attendee List
        attendee_list = []
        for email,contact in ds_dict.values():
            attendee_list.append({'email':email.strip()})
        attendee_list.append({'email':'sajna.vilangapurath@insofe.edu.in'})
                        
        ## Batch and Day filter         
        batch_id = re.findall(r'\d+', batch)[0]
        temp_schedule = schedule_dict[batch_id]
        entry = temp_schedule[temp_schedule['Date']==day]        
        assert (len(entry) == 1), "Multiple date entries"


        ## Mentor
        mentor_list = mentor_weekend_schedule.columns[(mentor_weekend_schedule[mentor_weekend_schedule['Date']==day]==batch).all()].tolist()
        if (len(mentor_list)==1):         
            mentor = mentor_list[0]
            if (mentor in contacts['Mapping'].tolist()) or (mentor in contacts['Name'].tolist()):
                mentor_dict = {mentor:contacts.loc[(contacts['Mapping']==mentor) |(contacts['Name']==mentor) ,['emailid','Contact#']].values[0]}
                mentor_details = ''.join(['%s \n%s - %s' % (key, value[0], value[1]) for (key, value) in mentor_dict.items()])
                mentor_email = mentor_dict[mentor][0]
            else:
                print("No contact info for ", mentor)
                mentor = ""
                mentor_dict = ""
                mentor_details = ""
                mentor_email = ""
        elif len(mentor_list)>1:
            raise("Multiple mentor allocations for batch ",batch)
        else:
            print("No mentor")
            mentor = ""
            mentor_dict = ""
            mentor_details = ""
            mentor_email = ""

        ## Module
        module_id = entry.loc[:,"Course #"].values[0]
        module_text = entry.iloc[:,4].values[0]
        cute = entry.loc[:,"CUTe"].values[0]
        if module_id in ['LAB DAY'] and cute not in [None, '', ' ']:
            topic = "LAB DAY & CUTe - " + module_text
            ds_attendee_type = "F"
            print("CUTe")
        elif module_id in ['LAB DAY','CSE 7323c','CSE 7212c']:
            topic = "LAB DAY - " + module_text
            ds_attendee_type = "F"
            print("LAB DAY - ",module_id)
        elif module_id == "MiTH":
            topic = "MiTH"
            ds_attendee_type = "F"
            mentor_attendee_type = "F"
            print("MiTH")
        elif module_id == "CSE 9099":
            topic = "PhD"
            ds_attendee_type = "F"
            mentor_attendee_type = "F"
            print("PhD")
        else:
            topic = entry.iloc[:,8].values[0] if entry.iloc[:,8].values[0] not in [None,""," "] else "" 
            ds_attendee_type = "A"
            mentor_attendee_type = "M"
        print("\n")

        module = (" - ".join([module_id,entry.iloc[:,4].values[0]])) if module_id not in ['LAB DAY', 'MiTH', 'CSE 9099'] else entry.iloc[:,4].values[0]
        rote = entry.loc[:,"ROTe"].values[0]
        if rote not in [None,""," "]:
            rote_ref = re.findall(r'Day \d+',rote)[0]
            rote = re.findall(r'Day \d+ Topics',rote)[0]
            rote_topics = temp_schedule[temp_schedule['Day']==rote_ref].iloc[:,8].values[0]
            if rote_topics in [None,' ']:
                rote_topics = ''
        else:
            rote_ref = ""
            rote_topics = ""
            
        disclaimer = "\n\n*This notification system is in beta phase. If you find anything suspicious or incorrect please write back at sagar.patel@insofe.edu.in\n Feedbacks welcome!!"
        
        note = ""
## Adding note for a specific day
# =============================================================================
#         if(day.weekday() == 6):
#             note = "<b>NOTE - The roads to INSOFE-HYD are schedule to be blocked till 10AM for Hyderabad Marathon.\nThe classes are rescheduled to start at 11AM</b>"
#         else:
#             note = ""
#         
# =============================================================================
        
        if mentor not in [None,'',' ']:
            mentor_summary = (topic + " @ " if topic not in [None,""," "] else "") + batch
            mentor_description = "\n".join([note if note not in [None,""," "] else '', "\n\n<b>Batch: </b>" + batch, "<b>Date: </b>" + day.strftime('%d-%b'),
                                            "\n<b>Module: </b>" + module,
                                            ("<b>Topics: </b>" + topic if topic not in [None,""," "] else ""),
                                            "\n<b>Data Scientists</b>" , ds_contacts, disclaimer])
            if run=='send':
                mentor_id = create_event(service=google_service,summary=mentor_summary, description=mentor_description ,dt = day, attendee_type= mentor_attendee_type, attendees=[{'email':mentor_email}])
            else:
                mentor_id = 0
        else:
            mentor_id = 'No mentor'
            
        ds_summary = (topic + " @ " if topic not in [None,""," "] else "") + batch
        ds_description = "\n".join([note if note not in [None,""," "] else '', "\n\n<b>Batch: </b>" + batch, "<b>Date: </b>" + day.strftime('%d-%b'),
                                    ("\n<b>ROTe: </b>" + rote if rote not in [None,""," "] else ''),
                                    ("<b>ROTe Topics: </b>" + rote_topics if rote_topics not in [None,""," "] else ''),
                                    "\n<b>Module: </b>" + module,
                                    ("<b>Topics to be taught: </b>" + topic if topic not in [None,""," "] else ''),
                                    ("\n<b>Mentor: </b>" + mentor_details  if mentor not in ["Data Scientists","", " "] else ''),
                                    "\n<b>Data Scientists</b>" , ds_contacts, disclaimer])
        if run=='send':   
            ds_id = create_event(service=google_service,summary=ds_summary, description=ds_description ,dt = day, attendee_type= ds_attendee_type, attendees=attendee_list) 
        else:
            ds_id = 0

        event_frame = event_frame.append({'dsEventId':ds_id,'mentorEventId':mentor_id,'Batch':batch, 'Date':day,
                                          'Module':module,'Topic':topic,'Mentor':mentor, 'MentorEventType':mentor_attendee_type , 'DSEventType':ds_attendee_type,
                                          'MentorEmail':mentor_email,'AttendeeList':attendee_list,
                                          'MentorDetails':mentor_details,'DSContact':ds_contacts.replace('\n', " || "),
                                          'ROTe':rote,'ROTeDay':rote_ref,'ROTeTopics':rote_topics}, ignore_index=True)        

    
event_frame.to_csv("notifications_archive/" + day.strftime('%d-%b') + "_Notifications.csv")
