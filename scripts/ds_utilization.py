#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Sep 25 11:46:43 2018

@author: sagar
"""
import sys
sys.path.insert(0, '/home/sagar/Dropbox/Scheduler/scripts/')
from openpyxl import load_workbook
import pandas as pd
from datetime import date
from aux_functions import date_range, next_dates, this_weekend, read_sheet
import re

scheduling_matrix = load_workbook("data/Faculty-Scientist_Scheduling_Matrix.xlsx",data_only=True)

ds_schedule = read_sheet(scheduling_matrix,'DS_PGP2018')
ds_schedule = ds_schedule.replace({"":None," ":None,"  ":None,"   ":None})
ds_schedule['Date'] = pd.to_datetime(ds_schedule['Date']).dt.date

melted_ds_schedule = pd.melt(ds_schedule,id_vars=['Date'],var_name="DataScientist",value_name="Allocation")
melted_ds_schedule['Type']=""
melted_ds_schedule['Hours']=0
melted_ds_schedule.loc[melted_ds_schedule['Allocation'].notnull(),'Type']="Corporate"
melted_ds_schedule.loc[melted_ds_schedule['Allocation'].notnull(),'Hours']=4
melted_ds_schedule.loc[melted_ds_schedule['Allocation'].str.contains("pgp",na=False,case=False).tolist(),'Type']="Academic"
melted_ds_schedule.loc[melted_ds_schedule['Allocation'].str.contains("pgp|cpee|Batch",na=False,case=False).tolist(),'Hours']=4
melted_ds_schedule.loc[melted_ds_schedule['Allocation'].str.contains("Rennes",na=False,case=False).tolist(),'Hours']=2
melted_ds_schedule.loc[melted_ds_schedule['Allocation'].str.contains("info",na=False,case=False).tolist(),'Type']="Other"
melted_ds_schedule.loc[melted_ds_schedule['Allocation'].str.contains("meet",na=False,case=False).tolist(),'Type']="Other"
melted_ds_schedule['Binary'] = [0 if x==None else 1 for x in melted_ds_schedule['Allocation']]
#melted_ds_schedule = melted_ds_schedule.fillna('')                           
#melted_ds_schedule['1W'] = melted_ds_schedule.groupby('DataScientist')['Binary'].rolling(7).sum().reset_index(0,drop=True)





program_schedule = load_workbook("data/Program Schedule.xlsx",data_only=True)
#program_schedule = load_workbook("data/ProgramSchedule_Batches 09-41.xlsx",data_only=True)

ps_schema = ["Week Day","Day","Date","Course #","Module","Mentor","ROTe","CUTe","Topics"]

           
# List of program schedules - 1 dataframe per batch                  
schedule_dict = {}
for sheet_name in program_schedule.sheetnames:
    sheet = read_sheet(program_schedule,sheet_name)
    if sheet.shape[1] >= 9:
        sheet = sheet.iloc[:,0:9]
        if False in set(ps_schema == sheet.columns):
            print('Column schema mismatch in sheet = ' + sheet_name)
            print(sheet.columns)
        else:
            schedule_dict[re.findall(r'\d+', sheet_name)[0]] = sheet
keys = schedule_dict.keys()


for idx, row in melted_ds_schedule.iterrows():
    if row['Allocation'] not in [None,'',' ']:
        if len(re.findall(r'\d+', row['Allocation'])) > 0:
            batch_lookup = re.findall(r'\d+', row['Allocation'])[0]
            if batch_lookup in keys:
                date_lookup = row['Date']
                ds_lookup = row['DataScientist']
                batch_lookup = re.findall(r'\d+', row['Allocation'])[0]
                temp_sheet = schedule_dict[batch_lookup]
                module = temp_sheet.loc[temp_sheet['Date'] == date_lookup, "Course #"].tolist()
                if len(module)>0:
                    if module[0] in ['LAB DAY','CSE 7323c','CSE 7212c','CSE 9099']:
                        melted_ds_schedule.loc[idx,'Hours'] = 8


#melted_copy = melted_ds_schedule.copy()
#melted_copy['Date'] = pd.DatetimeIndex(melted_copy['Date'])
#melted_copy.set_index(['Date'],inplace=True)
#melted_copy.groupby(["DataScientist",pd.Grouper(freq="M")],as_index=False).agg({'Hours':'sum'})
#
#melted_ds_schedule.groupby("") 

melted_ds_schedule['4W'] = melted_ds_schedule.groupby('DataScientist')['Binary'].rolling(28).sum().reset_index(0,drop=True)
melted_ds_schedule.to_csv('data/melted_ds_schedule.csv')
