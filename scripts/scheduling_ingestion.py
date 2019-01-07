from openpyxl import load_workbook
import pandas as pd
from datetime import date
from aux_functions import date_range, next_dates, this_weekend, read_sheet
import re


ps_schema = ["Week Day","Day","Date","Course #","Topics","Mentor","ROTe","CUTe","Topics"]
program_schedule = load_workbook("data/Program Schedule.xlsx",data_only=True)
           
                  
schedule_list = []
for sheet_name in program_schedule.sheetnames:
    sheet = read_sheet(program_schedule,sheet_name)
    if sheet.shape[1] >= 9:
        sheet = sheet.iloc[:,0:9]
        if False in set(ps_schema == sheet.columns):
            raise Exception('Column schema mismatch in sheet = ' + sheet_name)
        else:
            schedule_list.append(sheet)



#==============================================================================
# ## Mentor
#==============================================================================
scheduling_matrix = load_workbook("data/Faculty-Scientist_Scheduling_Matrix.xlsx",data_only=True)
mentor_schedule = read_sheet(scheduling_matrix,'Mentor_PGP2018')
mentor_schedule = mentor_schedule.replace({"":None," ":None,"  ":None,"   ":None})
mentor_schedule['Date'] = pd.to_datetime(mentor_schedule['Date']).dt.date
melted_mentor_schedule = pd.melt(mentor_schedule,id_vars=['Date'],var_name="Mentor",value_name="Allocation")
melted_mentor_schedule['Type']=""
melted_mentor_schedule.loc[melted_mentor_schedule['Allocation'].notnull(),'Type']="Corporate"
melted_mentor_schedule.loc[melted_mentor_schedule['Allocation'].str.contains("pgp|cpee|Complete|Batch",na=False,case=False).tolist(),'Type']="Academic"
#melted_mentor_schedule.loc[melted_mentor_schedule['Allocation'].str.contains("cpee",na=False,case=False).tolist(),'Type']="Academic"
melted_mentor_schedule.loc[melted_mentor_schedule['Allocation'].str.contains("info|meet",na=False,case=False).tolist(),'Type']="Other"
#melted_mentor_schedule.loc[melted_mentor_schedule['Allocation'].str.contains("batch",na=False,case=False).tolist(),'Type']="Other"
#melted_mentor_schedule.loc[melted_mentor_schedule['Allocation'].str.contains("meet",na=False,case=False).tolist(),'Type']="Other"
melted_mentor_schedule['Binary'] = [1 if x in ["Corporate","Academic"] else 0 for x in melted_mentor_schedule['Type']]
#melted_mentor_schedule['Binary'] = [0 if x==None else 1 for x in melted_mentor_schedule['Allocation']]
#melted_mentor_schedule = melted_mentor_schedule.fillna('')                           
#melted_mentor_schedule['1W'] = melted_mentor_schedule.groupby('Mentor')['Binary'].rolling(7).sum().reset_index(0,drop=True)
melted_mentor_schedule['4W'] = melted_mentor_schedule.groupby('Mentor')['Binary'].rolling(28).sum().reset_index(0,drop=True)
melted_mentor_schedule.to_csv('data/melted_mentor_schedule.csv')

#==============================================================================
# ## Testing
# #test = melted_mentor_schedule[melted_mentor_schedule['Binary']==1]
# #print(test['Allocation'].unique())
# #test = melted_mentor_schedule[(melted_mentor_schedule['Allocation'].isnull() & melted_mentor_schedule['Binary']==1)]
# #melted_mentor_schedule['Allocation'].isnull() = melted_mentor_schedule[melted_mentor_schedule['Allocation'].notnull()]
# 
#==============================================================================

#==============================================================================
# ## DS
#==============================================================================
ds_schedule = read_sheet(scheduling_matrix,'DS_PGP2018')
ds_schedule = ds_schedule.replace({"":None," ":None,"  ":None,"   ":None})
ds_schedule['Date'] = pd.to_datetime(ds_schedule['Date']).dt.date

melted_ds_schedule = pd.melt(ds_schedule,id_vars=['Date'],var_name="DataScientist",value_name="Allocation")
melted_ds_schedule['Type']=""
melted_ds_schedule.loc[melted_ds_schedule['Allocation'].notnull(),'Type']="Corporate"
melted_ds_schedule.loc[melted_ds_schedule['Allocation'].str.contains("pgp",na=False,case=False).tolist(),'Type']="Academic"
melted_ds_schedule.loc[melted_ds_schedule['Allocation'].str.contains("cpee",na=False,case=False).tolist(),'Type']="Academic"
melted_ds_schedule.loc[melted_ds_schedule['Allocation'].str.contains("info",na=False,case=False).tolist(),'Type']="Other"
melted_ds_schedule.loc[melted_ds_schedule['Allocation'].str.contains("meet",na=False,case=False).tolist(),'Type']="Other"
melted_ds_schedule['Binary'] = [0 if x==None else 1 for x in melted_ds_schedule['Allocation']]
#melted_ds_schedule = melted_ds_schedule.fillna('')                           
#melted_ds_schedule['1W'] = melted_ds_schedule.groupby('DataScientist')['Binary'].rolling(7).sum().reset_index(0,drop=True)
melted_ds_schedule['4W'] = melted_ds_schedule.groupby('DataScientist')['Binary'].rolling(28).sum().reset_index(0,drop=True)

melted_ds_schedule.to_csv('data/melted_ds_schedule.csv')


# =============================================================================
# ## Notifications
# =============================================================================

contacts = pd.read_csv("source_data/ContactList.csv")
contacts['Contact#'] = contacts['Contact#'].fillna(0).astype(int).astype(str)

google_service = instantiate()

p_index = [re.findall(r'\d+', x) for x in program_schedule.sheetnames]

#attendee_list = []
#for ds in contacts['emailid'][contacts['Mapping'].isin(ds_list)].values.tolist():
#    attendee_list.append({'email',ds.strip()})

weekend_schedule = ds_schedule[ds_schedule['Date'].isin(this_weekend())]

active_batches = []
for index, row in weekend_schedule.iloc[:,ds_schedule.columns != "Date"].iterrows():
    idx = row.isin(active_batches)
    active_batches.extend(set(row[~idx]))
active_batches.remove(None)

event_frame = pd.DataFrame(columns=['dsEventId','mentorEventId','Date','Batch','Module','Topic','Mentor', 'MentorEmail','AttendeeList','MentorDetails','DSContact','ROTe','ROTeDay','ROTeTopics'])
event_frame = pd.read_csv('/home/sagar/Dropbox/Scheduler/04-Aug_Notifications_bak.csv', index_col = 0)

for batch in active_batches:
    for day in this_weekend():
        ds_list = weekend_schedule.columns[(weekend_schedule[weekend_schedule['Date']==day]==batch).all()].tolist()        
        ds_dict = {}
        for ds in ds_list:
            ds_dict[ds] = contacts.loc[(contacts['Mapping']==ds) | (contacts['Name']==ds),['emailid','Contact#']].values[0]
        ds_contacts = '\n'.join(['%s - %s - %s' % (key, value[0], value[1]) for (key, value) in ds_dict.items()])

        attendee_list = []
        for email,contact in ds_dict.values():
            attendee_list.append({'email':email.strip()})
                                 
        batch_id = re.findall(r'\d+', batch)
        temp_schedule = schedule_list[p_index.index(batch_id)]
        entry = temp_schedule[temp_schedule['Date']==day]
        
        topic = entry.iloc[:,8].values[0] if entry.iloc[:,8].values[0] not in [None,""," "] else "" 

        mentor = entry.loc[:,"Mentor"].values[0] if entry.loc[:,"Mentor"].values[0] not in ['Data Scientists','Dr. Parag & Data Scientists'] else '' 
        
        if (mentor in contacts['Mapping'].unique()) or (mentor in contacts['Name'].unique()):
            mentor_dict = {mentor:contacts.loc[(contacts['Mapping']==mentor) |(contacts['Name']==mentor) ,['emailid','Contact#']].values[0]}
            mentor_details = ''.join(['%s \n%s - %s' % (key, value[0], value[1]) for (key, value) in mentor_dict.items()])
            mentor_email = mentor_dict[mentor][0]
        else:
            if mentor not in [None,""," "]:
                print("No contact info for ", mentor)
            mentor = ""
            mentor_dict = ""
            mentor_details = ""
            mentor_email = ""

# =============================================================================
#         if mentor not in [None,'',' ']:
#             mentor_dict = {mentor:contacts.loc[(contacts['Mapping']==mentor) |(contacts['Name']==mentor) ,['emailid','Contact#']].values[0]}
#             mentor_details = ''.join(['%s \n%s - %s' % (key, value[0], value[1]) for (key, value) in mentor_dict.items()])
#             mentor_email = mentor_dict[mentor][0]
#         else:
#             mentor_dict = ""
#             mentor_details = ""
#             mentor_email = ""
#             
# =============================================================================
        module_id = entry.loc[:,"Course #"].values[0]
        module = (" - ".join([module_id,entry.iloc[:,4].values[0]])) if module_id not in ['LAB DAY', 'MiTH', 'CSE 9099'] else entry.iloc[:,4].values[0]
        rote = entry.loc[:,"ROTe"].values[0]
        if rote not in [None,""," "]:
            rote_ref = re.findall(r'Day \d+',rote)[0]
            rote_topics = temp_schedule[temp_schedule['Day']==rote_ref].iloc[:,8].values[0]
            if rote_topics in [None,' ']:
                rote_topics = ''
        else:
            rote_ref = ""
            rote_topics = ""
            
        disclaimer = "\n\n*This notification system is in beta phase. If you find anything suspicious or incorrect please write back at sagar.patel@insofe.edu.in\n Feedbacks welcome!!"

        
        ds_summary = (topic + " @ " if topic not in [None,""," "] else "") + batch
        ds_description = "\n".join(["Batch = " + batch, "Date = " + day.strftime('%d-%b'),
                                    ("\nROTe - " + rote if rote not in [None,""," "] else ''),
                                    ("ROTe Topics - " + rote_topics if rote_topics not in [None,""," "] else ''),
                                    "\nModule = " + module,
                                    ("Topics to be taught = " + topic if topic not in [None,""," "] else ''),
                                    ("\nMentor - " + mentor_details  if mentor not in ["Data Scientists","", " "] else ''),
                                    "\nData Scientists" , ds_contacts, disclaimer])
#        create_event(service=google_service,summary=ds_summary, description=ds_description ,dt = day, attendee_type= "DS", attendees=attendee_list) 
#        ds_id = create_event(service=google_service,summary=ds_summary, description=ds_description ,dt = day, attendee_type= "DS", attendees=attendee_list) 
        ds_id = 0
#        print(ds_summary,ds_description)    
    
        if mentor not in [None,'',' ','Data Scientists','Dr. Parag & Data Scientists']:
            mentor_summary = (topic + " @ " if topic not in [None,""," "] else "") + batch
            mentor_description = "\n".join(["Batch = " + batch, "Date = " + day.strftime('%d-%b'),
                                            "\nModule = " + module,
                                            ("Topics = " + topic if topic not in [None,""," "] else ""),
                                            "\nData Scientists" , ds_contacts, disclaimer])
#            print(mentor_summary,mentor_description)
#        create_event(service=google_service,summary=mentor_summary, description=mentor_description ,dt = day, attendee_type= "Mentor", attendees=[{'email':'sagar.patel@insofe.edu.in'}])
#            mentor_id = create_event(service=google_service,summary=mentor_summary, description=mentor_description ,dt = day, attendee_type= "Mentor", attendees=[{'email':mentor_email}])
            mentor_id = 0
        else:
            mentor_id = ''
        event_frame = event_frame.append({'dsEventId':ds_id,'mentorEventId':mentor_id,'Batch':batch, 'Date':day,
                                          'Module':module,'Topic':topic,'Mentor':mentor,
                                          'MentorEmail':mentor_email,'AttendeeList':attendee_list,
                                          'MentorDetails':mentor_details,'DSContact':ds_contacts.replace('\n', " || "),
                                          'ROTe':rote,'ROTeDay':rote_ref,'ROTeTopics':rote_topics}, ignore_index=True)        

event_frame.to_csv(day.strftime('%d-%b') + "_Notifications.csv")
    



temp_frame = pd.DataFrame(columns=['dsEventId','mentorEventId','Date','Batch','Module','Topic','Mentor', 'MentorEmail','AttendeeList','MentorDetails','DSContact','ROTe','ROTeDay','ROTeTopics'])
temp_attendee_list = [{'email':"sagarpatel.exe@gmail.com"}]
temp_summary = (topic + " @ " if topic not in [None] else "") + "Batch" + batch
temp_description = "\n".join(["Batch = " + batch, "Module = " + module, ("Mentor = " + mentor if mentor not in ['Data Scientists'] else ''), ("Topics = " + topic if topic not in [None] else "")])
temp_description = "Test2 " + temp_description
temp_id = create_event(service=google_service,summary=temp_summary, description=temp_description ,dt = day, attendee_type= "DS", attendees=temp_attendee_list)
temp_frame = temp_frame.append({'dsEventId':temp_id,'mentorEventId':mentor_id,'Batch':batch, 'Date':day,
                                          'Module':module,'Topic':topic,'Mentor':mentor,
                                          'MentorEmail':mentor_email,'AttendeeList':temp_attendee_list,
                                          'MentorDetails':mentor_details,'DSContact':ds_contacts.replace('\n', " || "),
                                          'ROTe':rote,'ROTeDay':rote_ref,'ROTeTopics':rote_topics}, ignore_index=True)

    
    
    
## Deleting Event by ID    
google_service.events().delete(calendarId='primary', eventId='3o49tkvu9fpok3ledhh8gtim0k',sendNotifications=True).execute()

    
# =============================================================================
# ## INSTRUCTORS LIST
# =============================================================================


         
######## TESTING #########

sheet = program_schedule.get_sheet_by_name("Batch 43")
sheet = pd.DataFrame(sheet.values)
sheet = sheet.rename(columns=sheet.iloc[0])
sheet = sheet.drop(1)

columns_schema = ["Week Day","Day","Date","Course #","Topics","Mentor","ROTe","CUTe","Topics"]
if False in columns_schema == sheet.iloc[0,0:9]:
    raise ("Column schema mismatch at ",sheet_name)

######## TESTING END #####

weekends = date_range(date(2018, 6, 24),date(2018, 7, 24))

date.today()

for i in next_dates(date(2018,6,5),20):
    print(i.date().strftime('%d %b %y'))